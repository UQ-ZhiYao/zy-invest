"""
Excel parser service  v1.0.1
Fixed row parsing — handles sidebar navigation labels in column 0
"""
import io
import asyncio
from datetime import datetime, date
from typing import Optional
import openpyxl
import logging

logger = logging.getLogger(__name__)

SIDEBAR_LABELS = {
    'Portfolio Dashboard','Portfolio','Holdings','Dividend','Transaction',
    'Settlement','Others','Fund Info','Investors','Historical','Analysis',
    'Home / Portfolio / Transaction','Home / Portfolio / Settlement',
    'Home / Portfolio / Dividend','Home / Portfolio / Others',
    'Home / Fund Info / Investors','Home / Fund Info / Historical',
    'Home / Fund Info / Distributions','Home / Fund Info / Statement',
    'Home / Fund Info / Payment','Home / Portfolio / Holdings',
}

def clean_row(row):
    """Remove sidebar label from col 0 if present, shift data accordingly."""
    vals = list(row)
    if vals and str(vals[0]).strip() in SIDEBAR_LABELS:
        vals = vals[1:]  # drop sidebar label
    return vals

def safe_float(val) -> Optional[float]:
    try:
        if val is None or str(val).strip() in ('', '-', 'nan', 'None'):
            return None
        return float(val)
    except (ValueError, TypeError):
        return None

def safe_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None

def is_data_row(row):
    """Check if row contains actual data (has a date in col 0 after cleaning)."""
    cleaned = clean_row(row)
    if not cleaned:
        return False
    return safe_date(cleaned[0]) is not None

def parse_and_import(file_content: bytes, db) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    loop = asyncio.new_event_loop()
    try:
        results = loop.run_until_complete(_import_all(wb, db))
    finally:
        loop.close()
    return results

async def _import_all(wb, db) -> dict:
    return {
        "historical":    await _import_historical(wb, db),
        "investors":     await _import_investors(wb, db),
        "transactions":  await _import_transactions(wb, db),
        "settlement":    await _import_settlement(wb, db),
        "dividends":     await _import_dividends(wb, db),
        "distributions": await _import_distributions(wb, db),
        "others":        await _import_others(wb, db),
    }

async def _import_historical(wb, db) -> dict:
    if "Historical" not in wb.sheetnames:
        return {"skipped": True}
    ws = wb["Historical"]
    count = errors = 0
    for row in ws.iter_rows(min_row=12, values_only=True):
        r = clean_row(row)
        if not r or safe_date(r[0]) is None:
            continue
        try:
            d = safe_date(r[0])
            await db.execute(
                """INSERT INTO historical
                   (date,derivatives,securities,reits,bonds,money_market,
                    receivables,cash,mng_fees,ints_on_fees,loans,ints_on_loans,
                    capital,earnings,total_units,nta,is_locked,source)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,TRUE,'excel_import')
                   ON CONFLICT (date) DO UPDATE SET
                     nta=EXCLUDED.nta, securities=EXCLUDED.securities,
                     cash=EXCLUDED.cash, mng_fees=EXCLUDED.mng_fees,
                     total_units=EXCLUDED.total_units,
                     is_locked=TRUE, source='excel_import'
                   WHERE historical.is_locked=TRUE OR historical.source='excel_import'""",
                d,
                safe_float(r[1]) or 0,   # derivatives
                safe_float(r[2]) or 0,   # securities
                safe_float(r[3]) or 0,   # reits
                safe_float(r[4]) or 0,   # bonds
                safe_float(r[5]) or 0,   # mmf
                safe_float(r[6]) or 0,   # receivables
                safe_float(r[7]) or 0,   # cash
                safe_float(r[8]) or 0,   # mng_fees
                safe_float(r[9]) or 0,   # ints_on_fees
                safe_float(r[10]) or 0,  # loans
                safe_float(r[11]) or 0,  # ints_on_loans
                safe_float(r[12]) or 0,  # capital
                safe_float(r[13]) or 0,  # earnings
                safe_float(r[14]) or 0,  # total_units
                safe_float(r[15]) or 1,  # nta
            )
            count += 1
        except Exception as e:
            errors += 1
            if errors <= 3:
                logger.error(f"Historical row error: {e} — row: {r[:6]}")
    return {"imported": count, "errors": errors}

async def _import_investors(wb, db) -> dict:
    if "Investors" not in wb.sheetnames:
        return {"skipped": True}
    ws = wb["Investors"]
    count = 0
    for row in ws.iter_rows(min_row=11, values_only=True):
        r = clean_row(row)
        if not r or not r[0] or str(r[0]).strip() in ('Investors','Total',''):
            continue
        name = str(r[0]).strip()
        if not name or name in SIDEBAR_LABELS:
            continue
        try:
            await db.execute(
                """INSERT INTO investors
                   (name,units,vwap,total_costs,current_nta,market_value,unrealized_pl,realized_pl)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
                   ON CONFLICT DO NOTHING""",
                name,
                safe_float(r[1]) or 0,
                safe_float(r[2]),
                safe_float(r[3]) or 0,
                safe_float(r[4]),
                safe_float(r[5]),
                safe_float(r[6]),
                safe_float(r[7]) or 0,
            )
            count += 1
        except Exception as e:
            logger.error(f"Investor row error: {e}")
    return {"imported": count}

async def _import_transactions(wb, db) -> dict:
    if "Transaction" not in wb.sheetnames:
        return {"skipped": True}
    ws = wb["Transaction"]
    count = errors = 0
    # Get first investor as fallback
    inv = await db.fetchrow("SELECT id FROM investors LIMIT 1")
    inv_id = str(inv["id"]) if inv else None
    if not inv_id:
        return {"imported": 0, "errors": 0, "note": "No investors found — import investors first"}
    for row in ws.iter_rows(min_row=11, values_only=True):
        r = clean_row(row)
        if not r or safe_date(r[0]) is None:
            continue
        instrument = r[4] if len(r) > 4 else None
        if not instrument:
            continue
        try:
            units  = safe_float(r[5]) or 0
            price  = safe_float(r[6]) or 0
            amount = safe_float(r[7]) or (units * price)
            fees   = safe_float(r[8]) or 0
            net    = safe_float(r[9]) or (amount - fees)
            await db.execute(
                """INSERT INTO transactions
                   (date,investor_id,region,asset_class,sector,instrument,
                    units,price,amount,total_fees,net_amount,theme)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
                   ON CONFLICT DO NOTHING""",
                safe_date(r[0]),
                inv_id,
                str(r[1] or 'MY'),
                str(r[2] or 'Securities [H]'),
                str(r[3]) if r[3] else None,
                str(instrument),
                units, price, amount, fees, net,
                str(r[10]) if len(r) > 10 and r[10] else None,
            )
            count += 1
        except Exception as e:
            errors += 1
    return {"imported": count, "errors": errors}

async def _import_settlement(wb, db) -> dict:
    if "Settlement" not in wb.sheetnames:
        return {"skipped": True}
    ws = wb["Settlement"]
    count = errors = 0
    inv = await db.fetchrow("SELECT id FROM investors LIMIT 1")
    inv_id = str(inv["id"]) if inv else None
    if not inv_id:
        return {"imported": 0}
    for row in ws.iter_rows(min_row=11, values_only=True):
        r = clean_row(row)
        if not r or safe_date(r[0]) is None:
            continue
        instrument = r[4] if len(r) > 4 else None
        if not instrument:
            continue
        try:
            await db.execute(
                """INSERT INTO settlement
                   (date,investor_id,region,asset_class,sector,instrument,
                    units,bought_price,sale_price,profit_loss,return_pct,remark)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
                   ON CONFLICT DO NOTHING""",
                safe_date(r[0]), inv_id,
                str(r[1] or 'MY'), str(r[2] or ''),
                str(r[3]) if r[3] else None, str(instrument),
                safe_float(r[5]) or 0,
                safe_float(r[6]) or 0,
                safe_float(r[7]) or 0,
                safe_float(r[8]) or 0,
                safe_float(r[9]),
                str(r[10]) if len(r) > 10 and r[10] else None,
            )
            count += 1
        except Exception as e:
            errors += 1
    return {"imported": count, "errors": errors}

async def _import_dividends(wb, db) -> dict:
    if "Dividend" not in wb.sheetnames:
        return {"skipped": True}
    ws = wb["Dividend"]
    count = errors = 0
    inv = await db.fetchrow("SELECT id FROM investors LIMIT 1")
    inv_id = str(inv["id"]) if inv else None
    if not inv_id:
        return {"imported": 0}
    for row in ws.iter_rows(min_row=11, values_only=True):
        r = clean_row(row)
        ex_date = safe_date(r[1]) if len(r) > 1 else None
        if not ex_date:
            continue
        instrument = r[4] if len(r) > 4 else None
        if not instrument:
            continue
        try:
            await db.execute(
                """INSERT INTO dividends
                   (ann_date,ex_date,pmt_date,asset_class,instrument,
                    units,dps_sen,amount,entitlement,investor_id)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                   ON CONFLICT DO NOTHING""",
                safe_date(r[0]), ex_date,
                safe_date(r[2]) if len(r) > 2 else None,
                str(r[3]) if len(r) > 3 and r[3] else None,
                str(instrument),
                safe_float(r[5]) or 0,
                safe_float(r[6]) or 0,
                safe_float(r[7]) or 0,
                str(r[8]) if len(r) > 8 and r[8] else None,
                inv_id,
            )
            count += 1
        except Exception as e:
            errors += 1
    return {"imported": count, "errors": errors}

async def _import_distributions(wb, db) -> dict:
    if "Distributions" not in wb.sheetnames:
        return {"skipped": True}
    ws = wb["Distributions"]
    count = errors = 0
    for row in ws.iter_rows(min_row=11, values_only=True):
        r = clean_row(row)
        ex_date = safe_date(r[1]) if len(r) > 1 else None
        if not ex_date:
            continue
        title = r[4] if len(r) > 4 else None
        if not title or str(title).strip() in ('Title', ''):
            continue
        try:
            dist_type = 'interim' if 'Interim' in str(title) else \
                        'final'   if 'Final'   in str(title) else 'special'
            await db.execute(
                """INSERT INTO distributions
                   (ann_date,ex_date,pmt_date,financial_year,title,dist_type,
                    dps_sen,total_units,total_dividend,payout_ratio)
                   VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                   ON CONFLICT DO NOTHING""",
                safe_date(r[0]), ex_date,
                safe_date(r[2]) if len(r) > 2 else None,
                str(r[3]) if len(r) > 3 and r[3] else 'FY?',
                str(title), dist_type,
                safe_float(r[5]) or 0,
                safe_float(r[6]),
                safe_float(r[7]),
                safe_float(r[8]),
            )
            count += 1
        except Exception as e:
            errors += 1
    return {"imported": count, "errors": errors}

async def _import_others(wb, db) -> dict:
    if "Others" not in wb.sheetnames:
        return {"skipped": True}
    ws = wb["Others"]
    count = errors = 0
    for row in ws.iter_rows(min_row=11, values_only=True):
        r = clean_row(row)
        if not r or safe_date(r[0]) is None:
            continue
        try:
            await db.execute(
                """INSERT INTO others
                   (record_date,title,income_type,amount,platform,description)
                   VALUES ($1,$2,$3,$4,$5,$6)
                   ON CONFLICT DO NOTHING""",
                safe_date(r[0]),
                str(r[1]) if len(r) > 1 and r[1] else 'Other',
                str(r[2]) if len(r) > 2 and r[2] else 'Others',
                safe_float(r[3]) or 0,
                str(r[4]) if len(r) > 4 and r[4] else None,
                str(r[5]) if len(r) > 5 and r[5] else None,
            )
            count += 1
        except Exception as e:
            errors += 1
    return {"imported": count, "errors": errors}
